import json
import os
import sys

from openpyxl import load_workbook
from openpyxl.workbook import Workbook


def convert_from_excel_to_json(main_dir: str):
    wb = load_workbook(os.path.join(main_dir, 'test.xlsx'))
    awb = wb.active
    index = 1
    data = {}
    while True:
        c = awb.cell(1, index)
        if c.value is None:
            break
        if len(c.value) <= 0:
            break
        c2 = awb.cell(2, index)
        data[c.value] = int(c2.value)
        index += 1

    with open(os.path.join(main_dir, 'test.json'), 'w') as fp:
        json.dump(data, fp)
    os.remove(os.path.join(main_dir, 'test.xlsx'))


def convert_from_json_to_excel(main_dir: str):
    data = {}
    with open(os.path.join(main_dir, 'test.json')) as file:
        data = json.load(file)

    wb = Workbook()
    awb = wb.active
    index = 1
    for k, v in data.items():
        awb.cell(1, index).value = k
        awb.cell(2, index).value = v
        index += 1

    wb.save(os.path.join(main_dir, 'test.xlsx'))
    os.remove(os.path.join(main_dir, 'test.json'))


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Give 'x-j' or 'j-x'")
        exit(-1)

    conversion = sys.argv[1]
    main_dir = os.path.dirname(sys.argv[0])

    if conversion not in ('x-j', 'j-x'):
        print("Give 'x-j' or 'j-x'")
        exit(-1)
        
    if conversion == 'x-j':
        convert_from_excel_to_json(main_dir)
    if conversion == 'j-x':
        convert_from_json_to_excel(main_dir)